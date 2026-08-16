"""Dịch vụ phân tích cấu trúc tệp Excel (.xlsx)."""

import io
from datetime import date, datetime
from typing import Any

import openpyxl
from src.common.exceptions.error_codes import ErrorCode
from src.common.exceptions.infrastructure import InfrastructureException
from src.domain.data_source.file_parser import ParsedDataSourceResult
from src.domain.data_source.value_objects import (
    ColumnMetadata,
    SchemaMetadata,
    TableMetadata,
)

MAX_PREVIEW_ROWS = 5


class ExcelParser:
    """Công cụ đọc và phân tích cấu trúc schema tệp Excel."""

    def parse(self, file_bytes: bytes, filename: str, saved_path: str) -> ParsedDataSourceResult:
        """Đọc tệp Excel, bóc tách cấu trúc cột, kiểu dữ liệu và 5 dòng xem trước.

        Args:
            file_bytes: Dữ liệu nhị phân của tệp Excel.
            filename: Tên tệp gốc.
            saved_path: Đường dẫn lưu trữ tệp trên server.

        Returns:
            ParsedDataSourceResult: Cấu trúc schema và dữ liệu mẫu.
        """
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
            sheet_names = workbook.sheetnames
            if not sheet_names:
                raise ValueError("Tệp Excel không có trang tính (sheet) nào.")

            sheet = workbook[sheet_names[0]]
            return self._extract_sheet_data(sheet, filename, saved_path)
        except Exception as exc:
            raise InfrastructureException(
                code=ErrorCode.FILE_PARSING_ERROR,
                message=f"Không thể đọc hoặc phân tích cấu trúc tệp Excel: {filename}",
            ) from exc

    def _extract_sheet_data(self, sheet: Any, filename: str, saved_path: str) -> ParsedDataSourceResult:
        """Trích xuất dữ liệu của một sheet Excel."""
        rows_iter = sheet.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            return self._build_empty_result(filename, saved_path, sheet.title)

        column_names = [str(col).strip() if col is not None else f"Column_{i+1}" for i, col in enumerate(header_row)]
        all_rows: list[list[Any]] = [list(row) for row in rows_iter if any(v is not None for v in row)]

        preview_rows = self._build_preview_rows(column_names, all_rows[:MAX_PREVIEW_ROWS])
        columns_meta = self._infer_columns_metadata(column_names, all_rows)
        table_meta = TableMetadata(name=sheet.title or "Sheet1", columns=tuple(columns_meta))
        schema_meta = SchemaMetadata(tables=(table_meta,), relationships=())

        return ParsedDataSourceResult(
            name=filename,
            file_path=saved_path,
            file_type="EXCEL",
            schema_metadata=schema_meta,
            preview_rows=preview_rows,
            total_rows=len(all_rows),
        )

    def _infer_columns_metadata(self, col_names: list[str], data_rows: list[list[Any]]) -> list[ColumnMetadata]:
        """Suy luận kiểu dữ liệu và thuộc tính cho danh sách cột."""
        columns: list[ColumnMetadata] = []
        for idx, col_name in enumerate(col_names):
            values = [row[idx] for row in data_rows if idx < len(row) and row[idx] is not None]
            inferred_type = self._infer_single_column_type(values)
            columns.append(
                ColumnMetadata(
                    name=col_name,
                    data_type=inferred_type,
                    nullable=len(values) < len(data_rows),
                )
            )
        return columns

    def _infer_single_column_type(self, values: list[Any]) -> str:
        """Xác định kiểu dữ liệu từ danh sách giá trị thực tế của một cột."""
        if not values:
            return "TEXT"
        if all(isinstance(v, bool) for v in values):
            return "BOOLEAN"
        if all(isinstance(v, int) and not isinstance(v, bool) for v in values):
            return "NUMBER"
        if all(isinstance(v, (int, float)) and not isinstance(v, bool) for v in values):
            return "NUMBER"
        if all(isinstance(v, (datetime, date)) for v in values):
            return "DATETIME"
        if len(set(values)) <= 5 and len(values) >= 10:
            return "OPTION"
        return "TEXT"

    def _build_preview_rows(self, col_names: list[str], rows: list[list[Any]]) -> list[dict[str, Any]]:
        """Định dạng các dòng preview thành danh sách dict."""
        preview: list[dict[str, Any]] = []
        for row in rows:
            row_dict: dict[str, Any] = {}
            for idx, col in enumerate(col_names):
                val = row[idx] if idx < len(row) else None
                if isinstance(val, (datetime, date)):
                    row_dict[col] = val.isoformat()
                else:
                    row_dict[col] = val
            preview.append(row_dict)
        return preview

    def _build_empty_result(self, filename: str, saved_path: str, sheet_title: str) -> ParsedDataSourceResult:
        """Tạo kết quả rỗng khi sheet không có dữ liệu."""
        table_meta = TableMetadata(name=sheet_title or "Sheet1", columns=())
        schema_meta = SchemaMetadata(tables=(table_meta,), relationships=())
        return ParsedDataSourceResult(
            name=filename,
            file_path=saved_path,
            file_type="EXCEL",
            schema_metadata=schema_meta,
            preview_rows=[],
            total_rows=0,
        )
